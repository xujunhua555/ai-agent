"""
Email MCP Server based on FastMCP
提供邮件发送功能的 MCP 服务

注意：如果在 Windows 上遇到 asyncio 错误，请确保使用官方 Python 而非 Microsoft Store 版本
"""
import json
import os
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from typing import Optional, List
from pathlib import Path
from datetime import datetime, timedelta
from pydantic import BaseModel, Field

from mcp.server.fastmcp import FastMCP

# 创建 FastMCP 服务器实例
mcp = FastMCP("email-server")


@mcp.tool()
async def create_weekly_report(json_data_file: str, key_desc_file: str, output_file: Optional[str] = None) -> dict:
    """
    创建周报 Excel
    
    Args:
        json_data_file: JSON 数据文件路径
        key_desc_file: Key 描述文件路径
        output_file: 输出 Excel 文件路径（可选）
    
    Returns:
        生成结果，包含成功/失败状态和输出文件路径
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        
        print(f"正在读取数据文件：{json_data_file}")
        with open(json_data_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        print(f"共读取 {len(data)} 条数据")
        
        print(f"正在读取 Key 描述文件：{key_desc_file}")
        with open(key_desc_file, 'r', encoding='utf-8') as f:
            key_descriptions = json.load(f)
        
        print(f"共读取 {len(key_descriptions)} 个 Key 描述")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "周报"
        
        header_font = Font(bold=True, size=11, color='FFFFFF')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        
        cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        column_order = [
            'weekly_time', 'requirement_name', 'requirement_id', 'is_cosmic_metric',
            'product_manager', 'developer', 'total_task_man_days', 'used_task_man_days',
            'requirement_description', 'execution_role', 'standard_hours', 'actual_hours',
            'priority', 'requirement_status', 'other_notes', 'planned_completion_date',
            'actual_completion_date', 'work_content_progress', 'launch_date', 'is_external_support'
        ]
        
        print("正在写入表头...")
        for col_idx, key in enumerate(column_order, start=1):
            cn_desc = key_descriptions.get(key, key)
            cell = ws.cell(row=1, column=col_idx, value=cn_desc)
            cell.font = header_font
            cell.alignment = header_alignment
            cell.fill = header_fill
            cell.border = thin_border
        
        column_widths = {
            'weekly_time': 12, 'requirement_name': 35, 'requirement_id': 15,
            'is_cosmic_metric': 15, 'product_manager': 12, 'developer': 12,
            'total_task_man_days': 18, 'used_task_man_days': 22,
            'requirement_description': 40, 'execution_role': 12,
            'standard_hours': 15, 'actual_hours': 15, 'priority': 10,
            'requirement_status': 12, 'other_notes': 20,
            'planned_completion_date': 18, 'actual_completion_date': 18,
            'work_content_progress': 50, 'launch_date': 18, 'is_external_support': 15
        }
        
        for col_idx, key in enumerate(column_order, start=1):
            width = column_widths.get(key, 15)
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        
        print("正在写入数据...")
        for row_idx, item in enumerate(data, start=2):
            for col_idx, key in enumerate(column_order, start=1):
                value = item.get(key, '')
                if isinstance(value, (int, float)) and key in ['planned_completion_date', 'actual_completion_date', 'launch_date']:
                    try:
                        excel_epoch = datetime(1899, 12, 30)
                        actual_date = excel_epoch + timedelta(days=value)
                        value = actual_date.strftime('%Y-%m-%d')
                    except:
                        value = str(value)
                
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = cell_alignment
                cell.border = thin_border
                ws.row_dimensions[row_idx].height = 25
                
                if key in ['requirement_description', 'work_content_progress', 'other_notes']:
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                    ws.row_dimensions[row_idx].height = 60
        
        ws.auto_filter.ref = ws.dimensions
        
        print("正在添加汇总统计...")
        summary_ws = wb.create_sheet(title="汇总统计")
        
        developer_stats = {}
        for item in data:
            developer = item.get('developer', '未知')
            if developer not in developer_stats:
                developer_stats[developer] = {'count': 0, 'total_hours': 0, 'requirements': []}
            developer_stats[developer]['count'] += 1
            try:
                hours = float(item.get('actual_hours', 0))
                developer_stats[developer]['total_hours'] += hours
            except:
                pass
            developer_stats[developer]['requirements'].append(item.get('requirement_name', ''))
        
        summary_headers = ['开发人员', '需求数量', '总工时 (小时)', '需求列表']
        for col, header in enumerate(summary_headers, start=1):
            cell = summary_ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.alignment = header_alignment
            cell.fill = header_fill
        
        for row, (developer, stats) in enumerate(developer_stats.items(), start=2):
            summary_ws.cell(row=row, column=1, value=developer)
            summary_ws.cell(row=row, column=2, value=stats['count'])
            summary_ws.cell(row=row, column=3, value=stats['total_hours'])
            summary_ws.cell(row=row, column=4, value='\n'.join(stats['requirements']))
        
        summary_ws.column_dimensions['A'].width = 15
        summary_ws.column_dimensions['B'].width = 12
        summary_ws.column_dimensions['C'].width = 15
        summary_ws.column_dimensions['D'].width = 60
        
        if output_file:
            output_path = Path(output_file)
        else:
            output_path = Path(json_data_file).parent / "周报.xlsx"
        
        output_path = output_path.resolve()
        wb.save(output_path)
        print(f"Excel 已保存到：{output_path}")
        
        print("\n" + "=" * 60)
        print("周报汇总统计:")
        print("=" * 60)
        print(f"总需求数：{len(data)}")
        print(f"开发人员数：{len(developer_stats)}")
        print("\n各开发人员工作量:")
        for developer, stats in developer_stats.items():
            print(f"  - {developer}: {stats['count']} 个需求，{stats['total_hours']} 小时")
        
        return {
            "success": True,
            "message": f"周报已生成",
            "output_file": str(output_path),
            "total_requirements": len(data),
            "total_developers": len(developer_stats),
            "developer_stats": {
                dev: {"count": stats["count"], "total_hours": stats["total_hours"]}
                for dev, stats in developer_stats.items()
            }
        }
    
    except Exception as e:
        return {"success": False, "message": f"生成周报失败：{str(e)}"}


class EmailConfig(BaseModel):
    smtp_server: str = Field(..., description="SMTP 服务器地址")
    smtp_port: int = Field(..., description="SMTP 端口")
    username: str = Field(..., description="发件人邮箱账号")
    password: str = Field(..., description="邮箱密码或授权码")
    use_tls: bool = Field(default=True, description="是否使用 TLS 加密")
    sender_name: Optional[str] = Field(default=None, description="发件人名称")


@mcp.tool()
async def send_email(
    to: str, subject: str, body: str,
    name: Optional[str] = None, sender: Optional[str] = None, password: Optional[str] = None,
    html: bool = False, cc: Optional[str] = None, bcc: Optional[str] = None,
    attachments: Optional[List[str]] = None, config: Optional[EmailConfig] = None
) -> dict:
    try:
        if config is None:
            config = EmailConfig(
                smtp_server=os.getenv("SMTP_SERVER", "smtp.richinfo.cn"),
                smtp_port=int(os.getenv("SMTP_PORT", "465")),
                sender_name=name or os.getenv("SENDER_NAME"),
                username=sender or os.getenv("SMTP_USERNAME","xujunhua@richinfo.cn"),
                password=password or os.getenv("SMTP_PASSWORD","Xjh@770368"),
                use_tls=os.getenv("SMTP_USE_TLS", "false").lower() == "true"
            )
        
        if not config.username or not config.password:
            return {"success": False, "message": "缺少必要的 SMTP 配置：username 或 password"}
        
        msg = MIMEMultipart()
        msg["From"] = f"{config.sender_name or config.username} <{config.username}>"
        msg["To"] = to
        msg["Subject"] = subject
        
        if cc:
            msg["Cc"] = cc
        
        content_type = "html" if html else "plain"
        msg.attach(MIMEText(body, content_type, "utf-8"))
        
        if attachments:
            for file_path in attachments:
                if os.path.exists(file_path):
                    with open(file_path, "rb") as f:
                        attachment = MIMEBase("application", "octet-stream")
                        attachment.set_payload(f.read())
                        encoders.encode_base64(attachment)
                        filename = os.path.basename(file_path)
                        attachment.add_header("Content-Disposition", f'attachment; filename="{filename}"')
                        msg.attach(attachment)
                else:
                    return {"success": False, "message": f"附件文件不存在：{file_path}"}
        
        recipients = [to]
        if cc:
            recipients.extend([email.strip() for email in cc.split(",")])
        if bcc:
            recipients.extend([email.strip() for email in bcc.split(",")])
        
        try:
            if config.use_tls:
                server = smtplib.SMTP(config.smtp_server, config.smtp_port, timeout=10)
                server.starttls()
            else:
                server = smtplib.SMTP_SSL(config.smtp_server, config.smtp_port, timeout=10)
            
            server.login(config.username, config.password)
            server.sendmail(config.username, recipients, msg.as_string())
            server.quit()
        except smtplib.SMTPServerDisconnected as e:
            return {"success": False, "message": f"SMTP 服务器意外断开连接：{str(e)}"}
        
        return {"success": True, "message": f"邮件已成功发送至：{to}", "recipients": recipients}
    
    except smtplib.SMTPAuthenticationError as e:
        return {"success": False, "message": f"SMTP 认证失败：{str(e)}"}
    except smtplib.SMTPConnectError as e:
        return {"success": False, "message": f"无法连接到 SMTP 服务器：{str(e)}"}
    except Exception as e:
        return {"success": False, "message": f"发送邮件失败：{str(e)}"}


@mcp.tool()
async def send_simple_email(to: str, subject: str, body: str) -> dict:
    return await send_email(to=to, subject=subject, body=body, html=False)


@mcp.tool()
async def send_html_email(to: str, subject: str, html_content: str) -> dict:
    return await send_email(to=to, subject=subject, body=html_content, html=True)


@mcp.tool()
async def send_batch_emails(
    recipients: List[str], subject: str, body: str,
    html: bool = False, attachments: Optional[List[str]] = None,
    config: Optional[EmailConfig] = None
) -> dict:
    results = {"total": len(recipients), "success": 0, "failed": 0, "details": []}
    
    for recipient in recipients:
        result = await send_email(
            to=recipient, subject=subject, body=body, html=html,
            attachments=attachments, config=config
        )
        if result["success"]:
            results["success"] += 1
        else:
            results["failed"] += 1
        results["details"].append({"recipient": recipient, "status": "success" if result["success"] else "failed", "message": result["message"]})
    
    return results


@mcp.tool()
async def generate_and_send_email(
    json_data_file: str, key_desc_file: str, recipients: List[str],
    subject: str = "周报", body: str = "您好！周报内容见附件，请查收。",
    name: Optional[str] = None, sender: Optional[str] = None, password: Optional[str] = None,
    output_file: Optional[str] = None
) -> dict:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        
        print("\n" + "="*60)
        print("第一步：生成周报 Excel")
        print("="*60)
        
        with open(json_data_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        with open(key_desc_file, 'r', encoding='utf-8') as f:
            key_descriptions = json.load(f)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "周报"
        
        header_font = Font(bold=True, size=11, color='FFFFFF')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        column_order = [
            'weekly_time', 'requirement_name', 'requirement_id', 'is_cosmic_metric',
            'product_manager', 'developer', 'total_task_man_days', 'used_task_man_days',
            'requirement_description', 'execution_role', 'standard_hours', 'actual_hours',
            'priority', 'requirement_status', 'other_notes', 'planned_completion_date',
            'actual_completion_date', 'work_content_progress', 'launch_date', 'is_external_support'
        ]
        
        for col_idx, key in enumerate(column_order, start=1):
            cn_desc = key_descriptions.get(key, key)
            cell = ws.cell(row=1, column=col_idx, value=cn_desc)
            cell.font = header_font
            cell.alignment = header_alignment
            cell.fill = header_fill
            cell.border = thin_border
        
        column_widths = {
            'weekly_time': 12, 'requirement_name': 35, 'requirement_id': 15,
            'is_cosmic_metric': 15, 'product_manager': 12, 'developer': 12,
            'total_task_man_days': 18, 'used_task_man_days': 22,
            'requirement_description': 40, 'execution_role': 12,
            'standard_hours': 15, 'actual_hours': 15, 'priority': 10,
            'requirement_status': 12, 'other_notes': 20,
            'planned_completion_date': 18, 'actual_completion_date': 18,
            'work_content_progress': 50, 'launch_date': 18, 'is_external_support': 15
        }
        
        for col_idx, key in enumerate(column_order, start=1):
            width = column_widths.get(key, 15)
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        
        for row_idx, item in enumerate(data, start=2):
            for col_idx, key in enumerate(column_order, start=1):
                value = item.get(key, '')
                if isinstance(value, (int, float)) and key in ['planned_completion_date', 'actual_completion_date', 'launch_date']:
                    try:
                        excel_epoch = datetime(1899, 12, 30)
                        actual_date = excel_epoch + timedelta(days=value)
                        value = actual_date.strftime('%Y-%m-%d')
                    except:
                        value = str(value)
                
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = cell_alignment
                cell.border = thin_border
                ws.row_dimensions[row_idx].height = 25
                
                if key in ['requirement_description', 'work_content_progress', 'other_notes']:
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                    ws.row_dimensions[row_idx].height = 60
        
        ws.auto_filter.ref = ws.dimensions
        
        summary_ws = wb.create_sheet(title="汇总统计")
        developer_stats = {}
        for item in data:
            developer = item.get('developer', '未知')
            if developer not in developer_stats:
                developer_stats[developer] = {'count': 0, 'total_hours': 0, 'requirements': []}
            developer_stats[developer]['count'] += 1
            try:
                hours = float(item.get('actual_hours', 0))
                developer_stats[developer]['total_hours'] += hours
            except:
                pass
            developer_stats[developer]['requirements'].append(item.get('requirement_name', ''))
        
        summary_headers = ['开发人员', '需求数量', '总工时 (小时)', '需求列表']
        for col, header in enumerate(summary_headers, start=1):
            cell = summary_ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.alignment = header_alignment
            cell.fill = header_fill
        
        for row, (developer, stats) in enumerate(developer_stats.items(), start=2):
            summary_ws.cell(row=row, column=1, value=developer)
            summary_ws.cell(row=row, column=2, value=stats['count'])
            summary_ws.cell(row=row, column=3, value=stats['total_hours'])
            summary_ws.cell(row=row, column=4, value='\n'.join(stats['requirements']))
        
        summary_ws.column_dimensions['A'].width = 15
        summary_ws.column_dimensions['B'].width = 12
        summary_ws.column_dimensions['C'].width = 15
        summary_ws.column_dimensions['D'].width = 60
        
        if output_file:
            output_path = Path(output_file)
        else:
            output_path = Path(json_data_file).parent / "周报.xlsx"
        
        wb.save(output_path)
        print(f"Excel 已保存到：{output_path}")
        
        report_result = {
            "success": True, "message": "周报已生成",
            "output_file": str(output_path), "total_requirements": len(data),
            "total_developers": len(developer_stats),
            "developer_stats": {dev: {"count": stats["count"], "total_hours": stats["total_hours"]} for dev, stats in developer_stats.items()}
        }
        
        print("\n" + "="*60)
        print("第二步：批量发送邮件")
        print("="*60)
        
        batch_results = {"total": len(recipients), "success": 0, "failed": 0, "details": []}
        
        for recipient in recipients:
            result = await send_email(
                to=recipient, subject=subject, body=body, name=name, sender=sender, password=password,
                html=False, attachments=[str(output_path)]
            )
            if result["success"]:
                batch_results["success"] += 1
            else:
                batch_results["failed"] += 1
            batch_results["details"].append({"recipient": recipient, "status": "success" if result["success"] else "failed", "message": result["message"]})
        
        print("\n" + "="*60)
        print("最终结果")
        print("="*60)
        print(f"周报生成：成功")
        print(f"邮件发送：{batch_results['success']}/{batch_results['total']} 成功")
        
        return {
            "success": batch_results["success"] > 0,
            "message": f"周报已生成并发送成功 {batch_results['success']}/{batch_results['total']} 封邮件",
            "report_result": report_result,
            "batch_results": batch_results
        }
    
    except Exception as e:
        print(f"\n操作失败：{e}")
        import traceback
        traceback.print_exc()
        return {"success": False, "message": f"生成周报并发送邮件失败：{str(e)}"}


@mcp.tool()
async def get_leader_emails() -> dict:
    try:
        leader_emails = ["dongxiaobo@richinfo.cn", "shengdeyang@richinfo.cn", "xujunhua@richinfo.cn"]
        return {"success": True, "emails": leader_emails, "count": len(leader_emails)}
    except Exception as e:
        return {"success": False, "message": f"获取领导邮箱失败：{str(e)}"}


@mcp.tool()
async def test_smtp_connection(config: Optional[EmailConfig] = None) -> dict:
    try:
        if config is None:
            config = EmailConfig(
                smtp_server=os.getenv("SMTP_SERVER", "smtp.qq.com"),
                smtp_port=int(os.getenv("SMTP_PORT", "587")),
                username=os.getenv("SMTP_USERNAME"),
                password=os.getenv("SMTP_PASSWORD"),
                use_tls=os.getenv("SMTP_USE_TLS", "true").lower() == "true"
            )
        
        if not config.username or not config.password:
            return {"success": False, "message": "缺少 SMTP 配置 (username 或 password)"}
        
        if config.use_tls:
            server = smtplib.SMTP(config.smtp_server, config.smtp_port)
            server.starttls()
        else:
            server = smtplib.SMTP_SSL(config.smtp_server, config.smtp_port)
        
        server.login(config.username, config.password)
        server.quit()
        
        return {"success": True, "message": "SMTP 连接和认证成功", "server": f"{config.smtp_server}:{config.smtp_port}"}
    
    except Exception as e:
        return {"success": False, "message": f"SMTP 连接失败：{str(e)}"}


if __name__ == "__main__":
    mcp.run()
